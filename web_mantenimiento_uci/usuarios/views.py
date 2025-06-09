import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse,JsonResponse, HttpResponseForbidden
from django.template import loader
from django.contrib.auth import authenticate,login, logout
from django.contrib.auth.models import User,Group
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.utils import timezone
from .models import Incidencia,Material,Reporte,Notification, Personal, SolicitudSoporte,RespuestaSoporte,  MaterialIncidencia
from django.core.paginator import Paginator
from functools import wraps
from django.core import serializers
from django.contrib import messages
from django.db.models import Count
from django.db.models.functions import ExtractMonth
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, Alignment
from datetime import datetime
# Create your views here.

#Decorador personalizado para verificar los grupos a q mi usuario pertenece
#Se pasan por parametro los nombre de los grupos
def grupo_requerido(*nombres_grupos):
    """Decorador para verificar pertenencia a grupos"""
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            """
            Función principal que realiza la verificación:
            
            - Comprueba si el usuario autenticado pertenece a alguno de los grupos indicados.
            - Si pertenece: ejecuta la vista original.
            - Si NO pertenece: devuelve un error 403 (acceso prohibido).
            """
            if request.user.groups.filter(name__in=nombres_grupos).exists():
                return view_func(request, *args, **kwargs)
            return HttpResponseForbidden("No tienes permiso para acceder")
        return wrapper
    return decorator



#Vista para el login
def custom_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        
        context = {
            'credenciales':1,
        }
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('main/')
        else:
            return render(request, 'login.html', {'error': 'Credenciales inválidas'})
    else:
        return render(request, 'login.html')



@login_required
@grupo_requerido('administrador')
def usuarios(request):
    tableUsuario = User.objects.all()
    template = loader.get_template('all_usuarios.html')
    
    #Obtener palabra a buscar
    query = request.GET.get('q')
    
    #Condicion para buscar elementos en la tabla
    if query:
        tableUsuario = (tableUsuario.filter(username__icontains=query) | tableUsuario.filter(email__icontains=query) 
        | tableUsuario.filter(is_active__icontains=query) | tableUsuario.filter(last_login__icontains=query)
        | tableUsuario.filter(date_joined__icontains=query) 
        )
    #Paginacion
    paginator = Paginator(tableUsuario,10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    
    #Contenido mostrado en la pagina
    context = {
        'tableUsuario': tableUsuario,
        'page_obj': page_obj
    }
    
    if request.method == 'POST':
            action=request.POST.get('action')
                
            if action != 'delete':
                username = request.POST['username']
                name = request.POST['name']
                lastname = request.POST['lastname']  
                email = request.POST['email']
                password = request.POST['password']  
                if User.objects.filter(username=username).exists():
                    return HttpResponse("El usuario ya existe")
                else:
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=password,
                        first_name=name,
                        last_name=lastname
                    )
                    
                    return redirect('usuarios')
            else:
                ids=request.POST.getlist('ids')
                User.objects.filter(id__in=ids).delete()
                return redirect('usuarios')
                
    return HttpResponse(template.render(context,request))
    
@login_required
@grupo_requerido('administrador')
def seleccionar_usuario(request,item_id):
        user = get_object_or_404(User, id=item_id)  # Buscar el ítem en la base de datos
        
        if request.method == 'POST':
            username = request.POST.get('username')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            rol = request.POST.get('rol')
            
            #limpiar todos los roles 
            user.groups.clear()
            grupo = Group.objects.get(name=rol)
            
            # Actualizar los campos del usuario
            user.groups.add(grupo)
            user.username = username
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.save()  # Guardar los cambios en la base de datos
            
            if rol == "tecnico":
                trabajador = Personal.objects.filter(trabajador=user).exists()
                
                if not trabajador:
                    personal = Personal(
                        trabajador = user
                    )
                    personal.save()
            return redirect('usuarios')
            
        else:
            return render(request, 'editar_usuario.html', {'user': user})
        
@login_required  
def incidencias(request):
    template = loader.get_template('all_incidencias.html')
    
    if request.user.is_superuser:
        tableIncidencia = Incidencia.objects.all()
    elif request.user.groups.filter(name='tecnico').exists():
        personal = Personal.objects.get(trabajador=request.user)
        tableIncidencia = Incidencia.objects.filter(tecnico_asignado=personal)
    else:
        current_user = request.user
        tableIncidencia = Incidencia.objects.filter(usuario_reporte=current_user) 
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == "eliminar":
                ids=request.POST.getlist('ids')
                Incidencia.objects.filter(id__in=ids).delete()
                return redirect('incidencias')
    
    #Obtener elemento a buscar
    query = request.GET.get('q')
    
    #Condicion para buscar elementos en la tabla
    if query:
        tableIncidencia =( tableIncidencia.filter(tipo__icontains=query) | tableIncidencia.filter(prioridad__icontains=query)
        | tableIncidencia.filter(estado__icontains=query) | tableIncidencia.filter(fecha__icontains=query)
        | tableIncidencia.filter(ubicacion__icontains=query) | tableIncidencia.filter(descripcion__icontains=query)
        )
        
    #Paginacion
    paginator = Paginator(tableIncidencia,10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    tecnicos_disponibles = Personal.objects.filter(
        trabajador__groups__name='tecnico',
        incidencia__isnull=True
    )
    tecnicos = Personal.objects.all()
    
    context = {
        'tableIncidencia': tableIncidencia,
        'page_obj': page_obj,
        'tecnicos_disponibles': tecnicos_disponibles,
        'tecnicos':tecnicos,
        'materiales_disponibles': Material.objects.filter(cantidad__gt=0),
    }
    return HttpResponse(template.render(context,request))

@login_required
def reportar_incidencia(request):
    
    if request.method == "POST":
            
        
        tipo = request.POST.get('tipo_incidencia')
        prioridad = request.POST.get('prioridad')
        ubicacion = request.POST.get('ubicacion')
        descripcion = request.POST.get('descripcion')
        imagen = request.FILES.get('imagen')
        
        incidencia = Incidencia (
            tipo=tipo,
            prioridad=prioridad,
            ubicacion=ubicacion,
            descripcion=descripcion,
            fecha = timezone.now(),
            usuario_reporte=request.user,
            imagen=imagen
        )
        reporte = Reporte (
            fecha = timezone.now(),
            descripcion = " Reporte de Incidencia en el Servidor",
            estado ="pendiente",
            reporte_incidencia = incidencia
        )
        incidencia.save()
        reporte.save()
        return redirect('incidencias')
    
    return render(request , 'reportar_incidencia.html')


@login_required
def seleccionar_incidencia(request,item_id):
        incidencia = get_object_or_404(Incidencia, id=item_id)  # Buscar el ítem de incidencia en la base de datos
        reporte = get_object_or_404(Reporte,reporte_incidencia=incidencia) #Buscar el item de reporte para actualizar en la tabla de reportes
        
        if request.method == 'POST':
            tipo = request.POST.get('tipo')
            prioridad = request.POST.get('prioridad')
            estado = request.POST.get('estado')
            fecha = request.POST.get('fecha')
            ubicacion = request.POST.get('ubicacion')
            descripcion = request.POST.get('descripcion')
            
            # Actualizar los campos del usuario
            incidencia.tipo = tipo
            incidencia.prioridad = prioridad
            incidencia.fecha = fecha
            incidencia.ubicacion = ubicacion
            incidencia.descripcion = descripcion
           
            
            if estado is None:
                incidencia.estado = "pendiente"
                reporte.estado = "pendiente"
            else:
                incidencia.estado = estado
                reporte.estado = estado
            
            # Guardar los cambios en la base de datos
            incidencia.save()  
            reporte.save()
            return redirect('incidencias')
            
        else:
            return render(request, 'editar_incidencia.html', {'incidencia': incidencia})
    
@login_required
@grupo_requerido('almacenero','administrador')
def materiales(request):
    tableMaterial = Material.objects.all()
    
    #Paginacion
    paginator = Paginator(tableMaterial,10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'tableMaterial':tableMaterial,
        'page_obj':page_obj
    }
    
    if request.method == 'POST':
        
        action = request.POST.get('action')
        
        if action == "delete":
                ids=request.POST.getlist('ids')
                Material.objects.filter(id__in=ids).delete()
                return redirect('materiales')
        
        nombre = request.POST.get('username')
        tipo = request.POST.get('tipo_material')
        cantidad = request.POST.get('cantidad')
        
        material = Material(
            nombre = nombre,
            tipo = tipo,
            cantidad = cantidad
        )
        material.save()
        return redirect('materiales')
    
    return render(request,'all_materiales.html',context)

@grupo_requerido('almacenero','administrador')
def seleccionar_material(request,item_id):
    material= get_object_or_404(Material,id = item_id)
    
    if request.method == 'POST':
        nombre = request.POST.get('username')
        tipo = request.POST.get('tipo')
        cantidad = request.POST.get('cantidad')
        
        material.nombre = nombre
        material.tipo = tipo
        material.cantidad = cantidad
        material.save()
        return redirect('materiales')
    else:
        return render(request, 'editar_material.html', {'material': material})

@grupo_requerido('almacenero','administrador')
def reportes(request):
    
    
    fecha_inicial = request.GET.get('fechaInicio')
    fecha_final = request.GET.get('fechaFin')
    

    
    
        
    if fecha_inicial and fecha_final:
        tableReporte = Reporte.objects.filter( fecha__range=[fecha_inicial, fecha_final])
    else:
        tableReporte = Reporte.objects.all()
            
    totalReportes = tableReporte.count()
    reporte_resuelto = tableReporte.filter(estado='resuelto').count()
    reporte_pendiente = tableReporte.filter(estado='pendiente').count()
    reporte_enProceso = tableReporte.filter(estado = 'en_proceso').count()
    
    # Obtener todas las incidencias del año actual
    year = timezone.now().year
    incidencias_por_mes = (
        Incidencia.objects.filter(fecha__year=year)
        .annotate(mes=ExtractMonth('fecha'))
        .values('mes')
        .annotate(cantidad=Count('id'))
        .order_by('mes')
    )

    # Crear lista con 12 posiciones (uno por cada mes)
    data = [0] * 12  # Inicializar con ceros
    for item in incidencias_por_mes:
        mes_index = item['mes'] - 1  # Chart.js empieza en 0 = Enero
        data[mes_index] = item['cantidad']
    
    
    context = {
        'tableReporte' : tableReporte,
        'totalReportes' : totalReportes,
        'reporte_resuelto' : reporte_resuelto,
        'reporte_pendiente' : reporte_pendiente,
        'reporte_enProceso' : reporte_enProceso,
        'incidencias_data': data,
        'year': year
    }
    
    return render(request,'all_reportes.html',context)


def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def main(request):
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    unread_count = notifications.filter(is_read=False).count()
    return render(request, 'main.html', {
        'notifications': notifications,
        'unread_count': unread_count
    })



@require_POST
@login_required
def mark_as_read(request, notification_id):
    try:
        notification = Notification.objects.get(id=notification_id, user=request.user)
        notification.is_read = True
        notification.save()
        return JsonResponse({'status': 'success'})
    except Notification.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Notificación no encontrada'}, status=404)

def get_notifications(request):
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    data = serializers.serialize('json', notifications)
    return JsonResponse({'notifications': data}, safe=False)

@require_POST
@login_required
def delete_notification(request, notification_id):
    try:
        notification = Notification.objects.get(id=notification_id, user=request.user)
        notification.delete()
        return JsonResponse({'status': 'success'})
    except Notification.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Notificación no encontrada'}, status=404)
    

@login_required
@grupo_requerido('administrador')
def personal(request):
    tablePersonal = Personal.objects.all()
    template = loader.get_template('all_personal.html')
    
    #Obtener palabra a buscar
    query = request.GET.get('q')
    
    #Condicion para buscar elementos en la tabla
    if query:
        tablePersonal = (tablePersonal.filter( trabajador__username__icontains=query))
        
    #Paginacion
    paginator = Paginator(tablePersonal,10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    
    #Contenido mostrado en la pagina
    context = {
        'tablePersonal': tablePersonal,
        'page_obj': page_obj
    }
    
    return HttpResponse(template.render(context,request))

@grupo_requerido('administrador')
def asignar_tecnico(request):
    if request.method == 'POST':
        incidencia_id = request.POST.get('incidencia_id')
        tecnico_id = request.POST.get('tecnico_id')



        try:
            incidencia = Incidencia.objects.get(id=incidencia_id)
            tecnico = Personal.objects.get(id=tecnico_id)

            # Desasignar cualquier técnico anterior
            if incidencia.tecnico_asignado:
                antiguo = incidencia.tecnico_asignado
                antiguo.incidencia = None
                antiguo.save()

            # Asignar nuevo técnico
            incidencia.tecnico_asignado = tecnico
            incidencia.save()

            # Actualizar también el modelo Personal
            tecnico.incidencia = incidencia
            tecnico.save()

            messages.success(request, f"Técnico {tecnico.trabajador.username} asignado correctamente.")
        except (Incidencia.DoesNotExist, Personal.DoesNotExist) as e:
            messages.error(request, "Hubo un problema al asignar el técnico.")
    
    return redirect('incidencias')   

@grupo_requerido('administrador')
def quitar_tecnico(request, incidencia_id):
        incidencia = get_object_or_404(Incidencia, id=incidencia_id)
    
        tecnico = incidencia.tecnico_asignado
        if tecnico:
            incidencia.tecnico_asignado = None
            incidencia.save()

        
        #Opcional: también actualizar el modelo Personal si lo usas
        tecnico.incidencia = None
        tecnico.save()

        return redirect('incidencias')

@login_required
def solicitar_soporte(request):
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        descripcion = request.POST.get('descripcion')

        if tipo and descripcion:
            SolicitudSoporte.objects.create(
                usuario=request.user,
                tipo=tipo,
                descripcion=descripcion
            )
            messages.success(request, "Tu solicitud ha sido enviada correctamente.")
            return redirect('solicitar_soporte')

    # Obtener todas las solicitudes del usuario actual
    mis_solicitudes = SolicitudSoporte.objects.filter(usuario=request.user).order_by('-fecha_solicitud')

    # Paginación (opcional)
    paginator = Paginator(mis_solicitudes, 10)  # 10 por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'soporte/solicitar_soporte.html', {
        'page_obj': page_obj
    })

@login_required
@grupo_requerido('administrador')
def bandeja_entrada_soporte(request):
    if not request.user.groups.filter(name='administrador').exists():
        messages.error(request, "Acceso denegado.")
        return redirect('inicio')

    # Todas las solicitudes
    solicitudes = SolicitudSoporte.objects.all().order_by('-fecha_solicitud')

    return render(request, 'soporte/bandeja_entrada_soporte.html', {
        'solicitudes': solicitudes
    })
    

@login_required
def detalle_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudSoporte, id=solicitud_id)

    # Verificar permisos
    if solicitud.usuario != request.user and not request.user.groups.filter(name='administrador').exists():
        messages.error(request, "No tienes acceso a esta solicitud.")
        return redirect('solicitudes_soporte')

    if request.method == 'POST' and 'mensaje' in request.POST:
        mensaje = request.POST.get('mensaje', '').strip()
        if mensaje:
            RespuestaSoporte.objects.create(
                solicitud=solicitud,
                autor=request.user,
                mensaje=mensaje
            )
            return redirect('detalle_solicitud', solicitud_id=solicitud.id)

    return render(request, 'soporte/detalle_solicitud.html', {
        'solicitud': solicitud
    })
    
@login_required
def completar_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudSoporte, id=solicitud_id)

    if request.user == solicitud.usuario:
        solicitud.estado = 'resuelto'
        solicitud.save()
        messages.success(request, "La solicitud ha sido marcada como completada.")

    return redirect('detalle_solicitud', solicitud_id=solicitud.id)



@grupo_requerido('almacenero','administrador')
def exportar_dashboard(request):
    # Obtener datos del año actual
    year = timezone.now().year
    incidencias_por_mes = (
    Incidencia.objects.filter(fecha__year=year)
    .annotate(mes=ExtractMonth('fecha'))
    .annotate(cantidad=Count('id'))
    .values_list('mes', 'cantidad')
    .order_by('mes')
    )

    data_meses = [0] * 12
    for mes, cantidad in incidencias_por_mes:
        if 1 <= mes <= 12:
            data_meses[mes - 1] = cantidad

    # Contar estados
    pendientes = Incidencia.objects.filter(estado='pendiente').count()
    resueltos = Incidencia.objects.filter(estado='resuelto').count()
    en_proceso = Incidencia.objects.filter(estado='en_proceso').count()

    # Crear libro de Excel
    wb = openpyxl.Workbook()

    # Hoja 1: Datos de gráficos
    ws_estados = wb.active
    ws_estados.title = 'Estados'

    ws_estados.append(['Estado', 'Total'])
    ws_estados['A1'].font = Font(bold=True)
    ws_estados['B1'].font = Font(bold=True)

    ws_estados.append(['Pendiente', pendientes])
    ws_estados.append(['Resuelto', resueltos])
    ws_estados.append(['En Proceso', en_proceso])

    # Gráfico de pie (tarta)
    chart_estado = PieChart()
    labels = Reference(ws_estados, min_col=1, min_row=2, max_row=4)
    data = Reference(ws_estados, min_col=2, min_row=1, max_row=4)
    chart_estado.add_data(data, titles_from_data=True)
    chart_estado.set_categories(labels)
    chart_estado.title = "Reportes por Estado"
    chart_estado.style = 10  # Estilo de gráfico

    ws_estados.add_chart(chart_estado, "D2")

    # Hoja 2: Reportes por Mes
    ws_meses = wb.create_sheet(title="Por Mes")
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    ws_meses.append(['Mes', 'Reportes'])
    ws_meses['A1'].font = Font(bold=True)
    ws_meses['B1'].font = Font(bold=True)

    for i, (mes, valor) in enumerate(zip(meses, data_meses), start=2):
        ws_meses[f"A{i}"] = mes
        ws_meses[f"B{i}"] = valor

    # Gráfico de barras
    chart_mes = BarChart()
    chart_mes.type = "col"
    chart_mes.title = "Reportes por Mes"
    chart_mes.x_axis.title = "Mes"
    chart_mes.y_axis.title = "Incidencias"

    data_bar = Reference(ws_meses, min_col=2, min_row=1, max_row=13, max_col=2)
    cats = Reference(ws_meses, min_col=1, min_row=2, max_row=13)
    chart_mes.add_data(data_bar, titles_from_data=True)
    chart_mes.set_categories(cats)
    chart_mes.shape = 4  # Tamaño del gráfico

    ws_meses.add_chart(chart_mes, "D2")

    # Hoja 3: Tabla de reportes
    ws_reportes = wb.create_sheet(title="Listado")

    headers = ['ID', 'Fecha', 'Tipo', 'Descripción', 'Estado']
    ws_reportes.append(headers)
    for cell in ws_reportes["1"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    incidencias = Incidencia.objects.all().values_list('id', 'fecha', 'tipo', 'descripcion', 'estado')

    rows = []
    for incidencia in incidencias:
        id_incidencia, fecha, tipo, descripcion, estado = incidencia
        # Eliminar tzinfo si es necesario
        if fecha is not None and timezone.is_aware(fecha):
            fecha = timezone.make_naive(fecha)
        rows.append([id_incidencia, fecha, tipo, descripcion, estado])

    # Añadir al archivo Excel
    for row in rows:
        ws_reportes.append(row)

    # Guardar el archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Dashboard_Incidencias.xlsx'
    
    wb.save(response)
    return response



@grupo_requerido('almacenero','administrador')
def asignar_material(request):
    if request.method == 'POST':
        incidencia_id = request.POST.get('incidencia_id')
        material_id = request.POST.get('material')
        cantidad = int(request.POST.get('cantidad', 1))

        try:
            incidencia = get_object_or_404(Incidencia, id=incidencia_id)
            material = get_object_or_404(Material, id=material_id)

            # Verificar stock suficiente
            if material.cantidad < cantidad:
                messages.error(request, f"No hay suficiente stock de {material.nombre}")
                return redirect('detalle_incidencia', incidencia_id=incidencia_id)

            # Crear registro de uso
            MaterialIncidencia.objects.create(
                incidencia=incidencia,
                material=material,
                cantidad_usada=cantidad
            )

            # Restar del inventario
            material.cantidad -= cantidad
            material.save()

            messages.success(request, f"{cantidad} unidad(es) de '{material.nombre}' asignadas correctamente.")

        except Exception as e:
            messages.error(request, f"Error al asignar el material: {e}")

        return redirect('incidencias')  # Cambia por tu URL de lista de incidencias
    
    

def quitar_material(request):
    if request.method == 'POST':
        registro_id = request.POST.get('material_incidencia_id')

        try:
            registro = get_object_or_404(MaterialIncidencia, id=registro_id)
            cantidad_devuelta = registro.cantidad_usada
            material = registro.material

            # Devolver stock al material
            material.cantidad += cantidad_devuelta
            material.save()
            

            # Eliminar el registro
            registro.delete()
            
        except Exception as e:
            messages.error(request, f"Error al quitar el material: {e}")
    
    return redirect(request.META['HTTP_REFERER'])

